# Python program to
# access Excel files

# Import required library
# import xlwings as xw

import pandas as pd
# from openpyxl import load_workbook
import openpyxl
import xlrd

filename="TCO-DNPG-2307-00752.xls"
# wb = openpyxl.load_workbook(filename, data_only=True)
wb = xlrd.open_workbook(filename)

print(wb)
print(wb.sheets)
# for sheet in wb.worksheets:
#     print(sheet)

for sheet in wb.sheets():
    print(sheet.name, sheet.nrows, sheet.ncols)
    for row in range(0, sheet.nrows):
        header = sheet.row_values(row, start_colx=0, end_colx=None)
        print(header)


# xlsxfile = pd.ExcelFile(filename)
# for sh in xlsxfile.sheet_names:
#     dfxl = xlsxfile.parse(sh)
#     print(dfxl)

# Opening an excel file
# wb = xw.Book('example_001.xlsx')
# wb = xw.Book("TCO-DNPG-2307-00752.xlsx")
# inteminvdf = pd.read_excel('ItemInventory140723.xlsx', index_col="NameFromTaco", usecols=['FullName', 'NameFromTaco'])
inteminvdf = pd.read_excel('ItemInventory140723.xlsx', usecols=['FullName', 'NameFromTaco'])

print(inteminvdf)
print (inteminvdf.index)
# print(inteminvdf.loc['TS-W981', 'FullName'])
# Viewing available
# sheets in it
# wks = xw.sheets
# print("Available sheets :\n", wks)
# print(type(wks), f'count:{len(wks)}')
# # Selecting a sheet
# ws = wks[0]

# Selecting a value
# from the selected sheet
# val = ws.range("C1").value
# print("A value in sheet1 :", val)

# Automatic table
# detection from
# a cell
# automatic = ws.range("a10").expand().value
# print("Automatic Table :", automatic)

# linescount=wb.sheets[0].range('A' + str(wb.sheets[0].cells.last_cell.row)).end('up').row
# print(linescount)

data=[]
firstpage=True
boldataline=False
# rawdata=wks[0].range("c1").current_region.value
# print(rawdata)

# DNRefNum = rawdata[1][3].strip().split("-")[1][-1] + "".join(rawdata[1][3].strip().split("-")[-2:])
# TxnDate = rawdata[2][3].strip().split(" ")[-1]
DNRefNum = wb.sheets()[0].cell(1, 3).value.strip().split("-")[1][-1] + "".join(wb.sheets()[0].cell(1, 3).value.strip().split("-")[-2:])#[1][3].strip().split("-")[1][-1] + "".join(rawdata[1][3].strip().split("-")[-2:])
print(DNRefNum)
TxnDate = wb.sheets()[0].cell(2, 3).value.strip().split(" ")[-1]
Memo = DNRefNum
print(DNRefNum, TxnDate)
DeliveryNotedict={'DNRefNum':DNRefNum, 'TxnDate':TxnDate, 'Memo': Memo}
wks=wb.sheets()
for sheet in wks:
    # rawdata=ws.range("c1").current_region.value
    # print(rawdata)
    # print()
    irow=1
    movetonextpage=False
    for idxrow in range(0, sheet.nrows):
        row = sheet.row_values(idxrow, start_colx=0, end_colx=None)
        print(row)

    # for row in rawdata:
        # print(f'irow={irow}', row)
        irow+=1
        # print()
        if not None and row[0]=="Item No" :
            if firstpage:
                # print('firstpage')
                data.append(row)
                boldataline=True
                firstpage=False
                # continue
            else:
                # print('Not firstpage')
                # data.append(row)
                boldataline=True
                firstpage=False
        elif not movetonextpage:
            # print(boldataline, row)
            if boldataline:
                if row[0] is not None:
                    if not None and row[0].startswith('Jenis Barang'):
                        # print('ada Jenis Barang')
                        pass
                    elif row[0].startswith('Sub Total') and not None:
                        # print('sub total found')
                        boldataline=False
                        movetonextpage=True
                        continue
                    elif row[3] is not None:
                        if row[3].startswith('Hormat'):
                            # print('hormatkami found')
                            boldataline=False
                            movetonextpage=True
                            continue
                        else:
                            data.append(row)
                    else:
                        data.append(row)
                elif row[3] is not None:
                    if row[3].startswith('Hormat'):
                        boldataline=False
                        movetonextpage=True
                        # continue
                    else:
                        data.append(row)
                else:
                    data.append(row)

# print (f'data: {data}')
for idx, x in enumerate(data):
    for colidx, col in enumerate(x):
        if col == "":
            data[idx][colidx]=None
for idx, x in enumerate(data):
    print(idx, x)
print(f'len data={len(data)}')


newdata=[]

templist=[]
for idx, dt in enumerate(data):
    if dt[0] != 'Item No':
        # print(idx,'not item')
        if dt[0] is None:
            if dt[1]:
                templist[1]+=" " + dt[1]
            if dt[3]:
                templist.append(dt[3])
                templist[3]=templist[3].split("/")[0].strip()
                # templist[3]+=" " + dt[3]
            if dt[4]:
                templist[4]+=" " + dt[4]
            newdata.append(templist)

            templist=[]
        else:
            uom=dt[2].split(" ")[-1]
            # print(uom)
            dt.append(uom)
            dt[2]=int(dt[2].split(".")[0])
            templist=dt
            # print(templist)

for idx, x in enumerate(newdata):
    print(idx, x)
print(f'len newdata={len(newdata)}')


df=pd.DataFrame(newdata, columns=['Item No', 'Description', 'Quantity', 'No.SO/Ext.Doc.No.', 'LPN No.', 'UOM', 'Ext.Doc.No.'])#, columns=data[0]+"UOM")
# print(df)
# df=df.groupby(['No.SO/Ext.Doc.No.','Item No', 'UOM'])['Quantity'].sum().reset_index().sort_values(by=['Item No', 'No.SO/Ext.Doc.No.'])#.sort_values(by=['No.SO/Ext.Doc.No.'])
df=df.groupby(['Ext.Doc.No.', 'No.SO/Ext.Doc.No.','Item No', 'UOM'])['Quantity'].sum().reset_index().sort_values(by=['Ext.Doc.No.','No.SO/Ext.Doc.No.', 'Item No'])
df['NameFromTaco']=df['Item No']
print(df)
# df['FullName'] = inteminvdf.loc[df['Item No'],'FullName']
df=df.merge(inteminvdf)
print(df)
df=df.groupby(['Ext.Doc.No.', 'No.SO/Ext.Doc.No.','Item No', 'UOM', 'FullName'])['Quantity'].sum().reset_index().sort_values(by=['Ext.Doc.No.','No.SO/Ext.Doc.No.', 'Item No'])
print(df)
lst = df.to_dict('records')
# print(lst)
DeliveryNotedict['lines']=lst
print(DeliveryNotedict)